import gzip
import tkinter
from tkinter import filedialog
import os
import re
from colorama import Fore
import sys
import time
import openpyxl as xl

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, exceptions

#the following likely will be for the temporary code

from datetime import datetime, date, timedelta
import calendar

ascii_logo =  """      

     @@@@@@                                                                     
     @@@@@@                                                                     
    @@@@@@       @@@@@     @@@@  @@@@@     @@@@@   @@@@@@@@@   @@@@@@      @@@@ 
    @@@@@@       @@@@@     @@@@    @@@@@  @@@@@    @@@@@@@@@   @@@@@@@@    @@@@ 
   @@@@@@        @@@@@     @@@@     ,@@@@@@@@      @@@@        @@@@@@@@@   @@@@ 
   @@@@@@        @@@@@     @@@@       @@@@@@       @@@@@@@@@   @@@@  @@@@  @@@@ 
  @@@@@@         @@@@@     @@@@     &@@@@@@@@      @@@@        @@@@   @@@@ @@@@ 
  @@@@@@@@@@@@@   @@@@@@@@@@@@     @@@@@  @@@@@    @@@@@@@@@   @@@@    @@@@@@@@ 
 @@@@@@@@@@@@@      @@@@@@@@     @@@@@     @@@@@.  @@@@@@@@@   @@@@      @@@@@@ 

"""   
print(Fore.LIGHTMAGENTA_EX + ascii_logo)
print("===================Luxen's SMPE Top 10 Naton Log Data Scanner===================")
print("")
while True:
    try:
        print(Fore.LIGHTMAGENTA_EX + "Launching File Explorer...")
        print("")
        print("Please Select a folder containing the chat logs.")
        window = tkinter.Tk()
        window.withdraw()
        window.wm_attributes('-topmost', 1)
        logs_folder = filedialog.askdirectory()
        window.destroy()
        logs_list = os.listdir(logs_folder)
    except FileNotFoundError:
        print(Fore.WHITE + "Error: Log folder not given")
        askrelaunch = ""
        while askrelaunch.lower() != "y":
            askrelaunch = input("Relaunch File Explorer? (y/n)").lower()
            if askrelaunch == "n":    
                print("Exiting Program in 3 seconds.....")
                time.sleep(3)    
                sys.exit()
            elif askrelaunch != "y":
                print("Error: Invalid response: use 'y' for yes or 'n' for no")
        continue

    top_10_nation_list = {}
    for gzlog in logs_list:
        if gzlog.endswith('.gz') and (int(gzlog[0:4]) > 2021 or (int(gzlog[0:4]) == 2021 and int(gzlog[5:7]) >= 7)):
            print(Fore.BLUE + "Scanning {}".format(gzlog))
            try:
                log = gzip.open(logs_folder + "/" +gzlog, 'rt', encoding='cp437').read()
            except EOFError:
                print(Fore.RED + "Failed to Open File, Skipping")
                continue
            log = log.replace('§b', "")
            log = log.replace("§8", "")
            last_nation_inst = log.rfind('.[ Nations ].')
            if last_nation_inst != -1:
                list_nation_end = log.find('<<<',last_nation_inst)
                checknationtype = log[log.find('Nation Name - ',last_nation_inst)+14:log.find('Nation Name - ',last_nation_inst)+33] == "Number of Residents"
                server_ind_start = log.rfind("/INFO]: Connecting to ", 0, last_nation_inst) + len("/INFO]: Connecting to ")
                list_nation_page_num = log.find('<<<',last_nation_inst)+11
                #print(log[server_ind_start:server_ind_start+len("play.smpearth.com, 25565")], int(log[list_nation_page_num]), checknationtype)
                if int(log[list_nation_page_num]) == 1 and checknationtype and log[server_ind_start:server_ind_start+len("play.smpearth.com, 25565")] == "play.smpearth.com, 25565":
                    print(Fore.GREEN + "Found Top 10 Nations List")
                    nation_list_formatted = str(log[last_nation_inst:list_nation_end])
                    nation_list_formatted = nation_list_formatted.replace("[Render thread/INFO]: [System] [CHAT]", "")
                    nation_list_formatted = nation_list_formatted.replace("[ Nations ].__________________.oOo.", "")
                    nation_list_formatted = nation_list_formatted.replace("[main/INFO]: [CHAT]", "")
                    nation_list_formatted = nation_list_formatted.replace("[Render thread/INFO]: [CHAT]", "")

                    timestamp_list = list(re.finditer("(\[[0-9][0-9]\:[0-9][0-9]\:[0-9][0-9]\])", nation_list_formatted))
                    nations_res_count = {}
                    for i, timestamp in enumerate(timestamp_list):
                        if i != len(timestamp_list)-1:
                            next_timestamp = timestamp_list[i+1]
                            nation_info_inst = nation_list_formatted[timestamp.end(0)+ 1:next_timestamp.start(0)]
                            if "Nation Name - Number of Residents\n" in nation_info_inst:
                                continue
                            res_count = list(re.finditer("- \([0-9]+\)", nation_info_inst))[-1]
                            nations_res_count[nation_info_inst[1:res_count.start(0)-1]] = res_count.group(0)[3:-1]
                            
                    if gzlog[:10] in top_10_nation_list.keys():
                        top_10_nation_list.update({gzlog[:10] : nations_res_count})
                    else:
                        top_10_nation_list[gzlog[:10]] = nations_res_count

    #temp
    for date_, nation_info in top_10_nation_list.items():
        format_date = datetime.strptime(date_[:10], "%Y-%m-%d")
        print(Fore.YELLOW + "---- Date: {} {}, {} ----".format(format_date.day, calendar.month_name[format_date.month], format_date.year))
        print("Log # for date: {}".format(date_[11:]))
        iter_count = 1
        for nation, res_count in nation_info.items():
            print(Fore.GREEN +"#{} : ".format(iter_count) + Fore.LIGHTBLUE_EX + "Nation: " + Fore.LIGHTCYAN_EX + str(nation) + Fore.GREEN + " |" +
                  Fore.LIGHTBLUE_EX + " # of Residents: " + Fore.LIGHTCYAN_EX + str(res_count))
            iter_count += 1
        print("")

    print(Fore.GREEN + "Logs have been scanned. Data Organized")
    print("")

    print(Fore.LIGHTMAGENTA_EX + "1: Write data to existing excel file (MUST BE A FILE MADE BY THIS PROGRAM)")
    print("2: Create new excel file")
    askexcelwrite = ""
    while askexcelwrite.lower() not in ["1", "2", "one", "two"]:
        askexcelwrite = input(Fore.MAGENTA + "(1 or 2): ")
        askexcelwrite.lower()
        workbook = ""
        if askexcelwrite in ["1", "one"]:  
            while workbook == "":
                try:
                    print('')
                    print(Fore.LIGHTMAGENTA_EX + "Launching File Explorer...")
                    print('')
                    print("Select Existing File to write the data")
                    window = tkinter.Tk()
                    window.withdraw()
                    window.wm_attributes('-topmost', 1)
                    xl_file_path = filedialog.askopenfile(initialdir="~/Documents")
                    window.destroy()
                    workbook = xl.load_workbook(xl_file_path.name)                    
                except FileNotFoundError or AttributeError:
                    print(Fore.WHITE + "Error: File Not Selected")
                    print("Relaunching File Explorer in 3 seconds...")
                    time.sleep(3) 
                    continue
                except exceptions.InvalidFileException:
                    print(Fore.WHITE + "Error: Inavlid file type. must be '.xlsx'")
                    print("Relaunching File Explorer in 3 seconds...")
                    time.sleep(3) 
                    continue
        elif askexcelwrite in ["2", "two"] :
            while workbook == "":
                try:
                    print('')
                    print(Fore.LIGHTMAGENTA_EX + "Launching File Explorer...")
                    print('')
                    print("Select Desination for new Excel file.")
                    window = tkinter.Tk()
                    window.withdraw()
                    window.wm_attributes('-topmost', 1)
                    xl_file_path = filedialog.asksaveasfile(
                        filetypes=[("Excel Workbook","*.xlsx")],
                        defaultextension = [("Excel Workbook","*.xlsx")],
                        initialfile = "sheet",
                        initialdir="~/Documents"
                        )
                    window.destroy()
                    workbook = xl.Workbook()
                except FileNotFoundError:
                    print(Fore.WHITE + "Error: File Not Selected")
                    print("Relaunching File Explorer in 3 seconds...")
                    time.sleep(3) 
                    continue
        else:
            print("Error: Invalid response")
            continue

    sheet = workbook.active
    sheet['A1'].fill = PatternFill(start_color="6aa84f", end_color="6aa84f", fill_type = "solid")
    sheet['A1'].value = "Days Since Release (Date)"
    sheet['B1'].fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type = "solid")
    sheet['B1'].value = "#1"
    sheet['C1'].fill = PatternFill(start_color="b7b7b7", end_color="b7b7b7", fill_type = "solid")
    sheet['C1'].value = "#2"
    sheet['D1'].fill = PatternFill(start_color="a97b57", end_color="a97b57", fill_type = "solid")
    sheet['D1'].value = "#3"
    postion4_10 = [sheet.cell(row=1,column=i) for i in range(5,12)]
    for i, postion in enumerate(postion4_10):
        postion.value = "#" + str(i+4)
        postion.fill = PatternFill(start_color="00ff00", end_color="00ff00", fill_type = "solid")
    column_lengths = []
    for i in range(1,12):
        column_lengths.append(sheet.column_dimensions[get_column_letter(i)].width)
    
    for i in range((date.today() - date(2021, 7, 1)).days):
        i = i + 1
        j = (i*2)
        cell = sheet.cell(row=j,column=1)
        date_ = datetime(2021, 7, 1) + timedelta(days=i-1)
        cell.value = "Day {} ({} {}, {})".format(i, calendar.month_abbr[date_.month], date_.day, date_.year)
        cell.fill = PatternFill(start_color="b7e1cd", end_color="b7e1cd", fill_type = "solid")
        if len(cell.value) > column_lengths[0]:
            column_lengths[0] = len(cell.value)

        sheet.cell(row=j+1,column=1).value = "# of Residents"
        try:
            first_log_date = list(top_10_nation_list.keys())[0]
            first_log_date_formatted = datetime.strptime(first_log_date[:10], "%Y-%m-%d")
        except IndexError:
            first_log_date_formatted = None
        if date_ == first_log_date_formatted:
            column_iter = 2
            for nation, res_count in top_10_nation_list[first_log_date].items():

                nationcell = sheet.cell(row=j,column=column_iter)
                nationcell.value = str(nation)
                if len(nationcell.value) > column_lengths[column_iter-1]:
                    column_lengths[column_iter-1] = len(nationcell.value)

                res_count_cell = sheet.cell(row=j+1,column=column_iter)
                res_count_cell.value = str(res_count)
                if len(res_count_cell.value) > column_lengths[column_iter-1]:
                    column_lengths[column_iter-1] = len(res_count_cell.value)

                column_iter += 1
            top_10_nation_list.pop(first_log_date, None)
    for i, column_width in enumerate(column_lengths,1):
        sheet.column_dimensions[get_column_letter(i)].width = column_width

    workbook.save(filename = xl_file_path.name)
    workbook.close()
    print("")
    print(Fore.GREEN + "Data successfully transfered to spreadsheet")
    print("")
    asktorepeat = ""
    while asktorepeat.lower() != "y":
        asktorepeat = input(Fore.LIGHTMAGENTA_EX + "Open another folder? (y/n)")
        asktorepeat.lower()
        if asktorepeat == "n":    
            print("Exiting Program in 3 seconds.....")
            time.sleep(3)    
            sys.exit()
        elif asktorepeat != "y":
            print("Error: Invalid response: use 'y' for yes or 'n' for no")
    continue

        
